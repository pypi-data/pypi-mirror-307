from io import BytesIO
import re
from typing import Tuple, Union
from openpyxl import Workbook, load_workbook
from pandas import DataFrame, read_excel, read_csv
from sharepoint_utils.basic_utils import get_item, write_item
from sharepoint_utils.connection import SharePointConnection


def get_excel_as_df(
    sharepoint_connection: SharePointConnection,
    drive: str,
    file_path: str,
    sheet_name: str = None,
    subsite: str = None,
    header: Union[int, list] = 0,
    **kwargs,
) -> DataFrame:
    """
    Get a Pandas DataFrame from a SharePoint Excel file

    Parameters
    ----------
    * sharepoint_connection (SharePointConnection): connection to SharePoint
    * drive (str): name of drive
    * file_path (str): path to the excel file
    * sheet_name (str, optional): name of the excel sheet to use, defaults to None
    * subsite (str, optional): name of subsite where folder is located, defaults to None
    * header (int or list, optional): which row to use as column names, defaults to 0
    * kwargs (optional): any other keyword args to pass to Pandas read_excel

    Returns
    -------
    * Pandas DataFrame or dict[str, DataFrame] if sheet_name is not specified
    """
    excel_content = get_item(sharepoint_connection, drive, file_path, subsite=subsite)
    df = read_excel(
        excel_content, sheet_name=sheet_name, engine="openpyxl", header=header, **kwargs
    )
    if type(header) is list:
        # Replace column labels starting with "Unnamed:" with empty strings
        df.columns = [
            _replace_unnamed_vals_for_multiindex_cols(col_val) for col_val in df.columns
        ]
    return df


def _replace_unnamed_vals_for_multiindex_cols(col_val: Tuple) -> Tuple:
    return tuple([re.sub(r"^Unnamed:.*", "", val) for val in col_val])


def write_df_to_excel(
    sharepoint_connection: SharePointConnection,
    df: DataFrame,
    drive: str,
    folder_path: str,
    file_name: str,
    subsite: str = None,
    **kwargs,
):
    """
    Write a Pandas DataFrame to an Excel file on SharePoint

    Parameters
    ----------
    * sharepoint_connection (SharePointConnection): connection to SharePoint
    * drive (str): name of drive
    * file_path (str): path to the excel file
    * sheet_name (str, optional): name of the excel sheet to use, defaults to None
    * subsite (str, optional): name of subsite where folder is located, defaults to None
    * header (int or list, optional): which row to use as column names, defaults to 0
    * kwargs (optional): any other keyword args to pass to Pandas to_excel

    Returns
    -------
    * Pandas DataFrame
    """
    buffer = BytesIO()
    df.to_excel(buffer, **kwargs)
    data = buffer.getvalue()
    write_item(
        sharepoint_connection,
        drive,
        data,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_name,
        folder_path,
        subsite=subsite,
    )


def get_excel_as_workbook(
    sharepoint_connection: SharePointConnection,
    drive: str,
    file_path: str,
    subsite=None,
    **kwargs,
) -> Workbook:
    """
    Load a SharePoint Excel file as an OpenPyXL Workbook

    Parameters
    ----------
    * sharepoint_connection (SharePointConnection): connection to SharePoint
    * drive (str): name of drive
    * file_path (str): path to the excel file
    * subsite (str, optional): name of subsite where folder is located, defaults to None
    * kwargs (optional): any other keyword args to pass to OpenPyXL load_workbook

    Returns
    -------
    * OpenPyXL workbook
    """
    excel_content = get_item(sharepoint_connection, drive, file_path, subsite=subsite)
    return load_workbook(excel_content, **kwargs)


def write_workbook_to_excel(
    sharepoint_connection: SharePointConnection,
    workbook: Workbook,
    drive: str,
    folder_path: str,
    file_name: str,
    subsite=None,
):
    """
    Save an OpenPyXL Workbook to SharePoint as an Excel file

    Parameters
    ----------
    * sharepoint_connection (SharePointConnection): connection to SharePoint
    * workbook (OpenPyXL Workbook)
    * drive (str): name of drive
    * folder_path (str): where to save the Excel file
    * file_name (str): what to call the new Excel file
    * subsite (str, optional): name of subsite where folder is located, defaults to None

    Returns
    -------
    * None
    """
    buffer = BytesIO()
    workbook.save(buffer)
    data = buffer.getvalue()
    write_item(
        sharepoint_connection,
        drive,
        data,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        file_name,
        folder_path,
        subsite=subsite,
    )


def get_csv_as_df(
    sharepoint_connection: SharePointConnection,
    drive: str,
    file_path: str,
    subsite=None,
    **kwargs,
):
    """
    Get a Pandas DataFrame from a SharePoint CSV

    Parameters
    ----------
    * sharepoint_connection (SharePointConnection): connection to SharePoint
    * drive (str): name of drive
    * file_path (str): path to the CSV file
    * subsite (str, optional): name of subsite where file is located, defaults to None
    * kwargs (optional): any other keyword args to pass to Pandas read_csv

    Returns
    -------
    * Pandas DataFrame
    """
    csv_content = get_item(sharepoint_connection, drive, file_path, subsite=subsite)
    return read_csv(csv_content, **kwargs)


def write_df_to_csv(
    sharepoint_connection: SharePointConnection,
    df: DataFrame,
    drive: str,
    folder_path: str,
    file_name: str,
    subsite=None,
    **kwargs,
):
    """
    Save a Pandas DataFrame to SharePoint as a CSV

    Parameters
    ----------
    * sharepoint_connection (SharePointConnection): connection to SharePoint
    * df (Pandas DataFrame)
    * drive (str): name of drive
    * folder_path (str): where to save the CSV
    * file_name (str): what to call the new CSV
    * subsite (str, optional): name of subsite where file should be located, defaults to None
    * kwargs (optional): any other keyword args to pass to Pandas to_csv

    Returns
    -------
    * None
    """
    file_data = df.to_csv(index=False, **kwargs).encode("utf-8")
    write_item(
        sharepoint_connection,
        drive,
        file_data,
        "text/csv",
        file_name,
        folder_path,
        subsite=subsite,
    )
