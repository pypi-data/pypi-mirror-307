from io import BytesIO
from unittest.mock import MagicMock, call, patch
import pytest
from sharepoint_utils.spreadsheet_utils import (
    get_excel_as_df,
    get_excel_as_workbook,
    get_csv_as_df,
)
from openpyxl import Workbook
from pandas import DataFrame, MultiIndex


def _create_fake_excel_bytes(data):
    """Returns BytesIO object in an Excel format with some fake data"""
    wb = Workbook()
    sheet = wb.active
    for i, row in enumerate(data):
        sheet["A" + str(i + 1)] = row[0]
        sheet["B" + str(i + 1)] = row[1]

    buffer = BytesIO()
    wb.save(buffer)

    return buffer


base_data = [["country", "capitol"], ["US", "Washington DC"], ["Colombia", "Bogota"]]
base_csv = b"country, capitol\n US, Washington DC\n Colombia, Bogota"
mock_data_lst = [
    {
        "data": _create_fake_excel_bytes(base_data),
        "df": DataFrame(base_data),
        "header": 0,
        "multi_header": [],
        "dataType": "excel",
    },
    {
        "data": _create_fake_excel_bytes([["", ""]] + base_data),
        "df": DataFrame(base_data),
        "header": 1,
        "multi_header": [],
        "dataType": "excel",
    },
    {
        "data": _create_fake_excel_bytes([["A", "B"]] + base_data),
        "df": DataFrame(
            base_data,
            columns=MultiIndex.from_arrays([["A", "B"], ["country", "capitol"]]),
        ),
        "header": [0, 1],
        "multi_header": [("A", "country"), ("B", "capitol")],
        "dataType": "excel",
    },
    {
        "data": BytesIO(base_csv),
        "df": DataFrame(base_data),
        "header": 0,
        "multi_header": [],
        "dataType": "csv",
    },
]


@pytest.mark.parametrize(
    "mock_data",
    [
        row
        for row in mock_data_lst
        if len(row["multi_header"]) == 0 and row["dataType"] == "excel"
    ],
)
@patch("sharepoint_utils.spreadsheet_utils.get_item")
@patch("sharepoint_utils.spreadsheet_utils.read_excel")
def test_get_excel_as_df(mock_read_excel, mock_get_item, mock_connection, mock_data):
    mock_get_item.return_value = mock_data["data"]
    mock_read_excel.return_value = mock_data["df"]
    result = get_excel_as_df(
        mock_connection,
        "mock_drive",
        "mock_file_path",
        sheet_name=0,
        header=mock_data["header"],
    )
    assert isinstance(result, DataFrame), "Didn't get a dataframe back"
    mock_get_item.assert_called_once_with(
        mock_connection, "mock_drive", "mock_file_path", subsite=None
    )
    mock_read_excel.assert_called_once_with(
        mock_get_item.return_value,
        sheet_name=0,
        engine="openpyxl",
        header=mock_data["header"],
    )


@pytest.mark.parametrize(
    "mock_data", [row for row in mock_data_lst if len(row["multi_header"]) > 0]
)
@patch("sharepoint_utils.spreadsheet_utils.get_item")
@patch("sharepoint_utils.spreadsheet_utils.read_excel")
@patch("sharepoint_utils.spreadsheet_utils._replace_unnamed_vals_for_multiindex_cols")
def test_get_excel_as_df_multiheader(
    mock_replace_multiindex, mock_read_excel, mock_get_item, mock_connection, mock_data
):
    mock_get_item.return_value = mock_data["data"]
    mock_read_excel.return_value = mock_data["df"]
    mock_replace_multiindex.side_effect = mock_data["multi_header"]
    result = get_excel_as_df(
        mock_connection,
        "mock_drive",
        "mock_file_path",
        sheet_name=0,
        header=mock_data["header"],
    )
    assert isinstance(result, DataFrame), "Didn't get a dataframe back"
    mock_get_item.assert_called_once_with(
        mock_connection, "mock_drive", "mock_file_path", subsite=None
    )
    mock_read_excel.assert_called_once_with(
        mock_get_item.return_value,
        sheet_name=0,
        engine="openpyxl",
        header=mock_data["header"],
    )
    mock_replace_multiindex.assert_has_calls(
        [call(mock_data["multi_header"][0]), call(mock_data["multi_header"][1])]
    )


@patch("sharepoint_utils.spreadsheet_utils.get_item")
@patch("sharepoint_utils.spreadsheet_utils.load_workbook")
def test_get_excel_as_workbook(mock_load_workbook, mock_get_item, mock_connection):
    mock_get_item.return_value = mock_data_lst[0]["data"]
    mock_load_workbook.return_value = Workbook()
    result = get_excel_as_workbook(mock_connection, "mock_drive", "mock_file_path")
    assert isinstance(result, Workbook), "Didn't get back a workbook"
    mock_get_item.assert_called_once_with(
        mock_connection, "mock_drive", "mock_file_path", subsite=None
    )
    mock_load_workbook.assert_called_once_with(mock_get_item.return_value)


@pytest.mark.parametrize(
    "mock_data", [row for row in mock_data_lst if row["dataType"] == "csv"]
)
@patch("sharepoint_utils.spreadsheet_utils.get_item")
@patch("sharepoint_utils.spreadsheet_utils.read_csv")
def test_get_csv_as_df(mock_read_csv, mock_get_item, mock_connection, mock_data):
    mock_get_item.return_value = mock_data["data"]
    mock_read_csv.return_value = mock_data["df"]
    result = get_csv_as_df(mock_connection, "mock_drive", "mock_file_path")
    assert isinstance(result, DataFrame), "Didn't get back a dataframe"
    mock_get_item.assert_called_once_with(
        mock_connection, "mock_drive", "mock_file_path", subsite=None
    )
    mock_read_csv.assert_called_once_with(mock_get_item.return_value)
