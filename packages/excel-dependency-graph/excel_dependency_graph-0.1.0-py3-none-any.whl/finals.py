import openpyxl
import re
from typing import Dict, List, Set

class ExcelDependencyGraph:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=False)
        
        # Dictionary to store dependencies and dependents
        self.dependencies: Dict[str, Set[str]] = {}
        self.dependents: Dict[str, Set[str]] = {}
        
        # Pattern to match single cell references (e.g., A1) and ranges (e.g., A1:B2)
        self.formula_pattern = r'(\b[A-Z]+\$?\d+\b)'

    def parse_worksheet(self, sheet_name: str) -> None:
        """Parse a worksheet and populate dependencies and dependents dictionaries."""
        sheet = self.wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_coord = f"{sheet_name}!{cell.coordinate}"
                    formula = cell.value
                    dependencies = self._extract_dependencies(formula, sheet_name)
                    
                    # Initialize sets if cell_coord is new in the dictionaries
                    self.dependencies.setdefault(cell_coord, set())
                    self.dependents.setdefault(cell_coord, set())
                    
                    # For each dependency, update the dictionaries
                    for dep in dependencies:
                        self.dependencies[cell_coord].add(dep)
                        if dep not in self.dependents:
                            self.dependents[dep] = set()
                        self.dependents[dep].add(cell_coord)

    def _extract_dependencies(self, formula: str, sheet_name: str) -> Set[str]:
        """Extract cell dependencies from a formula using regex and return them as full sheet-cell refs."""
        matches = re.findall(self.formula_pattern, formula)
        dependencies = set()
        for match in matches:
            # Ensure each dependency is prefixed with the sheet name
            if '!' not in match:
                match = f"{sheet_name}!{match}"
            dependencies.add(match)
        return dependencies

    def build_dependency_graph(self) -> None:
        """Build the dependency graph by parsing each sheet in the workbook."""
        for sheet_name in self.wb.sheetnames:
            self.parse_worksheet(sheet_name)

    def get_dependencies(self, cell: str) -> List[str]:
        """Return a list of cells that the given cell depends on."""
        return list(self.dependencies.get(cell, []))

    def get_dependents(self, cell: str) -> List[str]:
        """Return a list of cells that depend on the given cell."""
        return list(self.dependents.get(cell, []))

    def print_graph(self) -> None:
        """Prints the dependency and dependent relations for all cells."""
        print("Dependencies:")
        for cell, deps in self.dependencies.items():
            print(f"{cell} depends on {deps}")
        
        print("\nDependents:")
        for cell, deps in self.dependents.items():
            print(f"{cell} is needed by {deps}")

# Example usage:
graph = ExcelDependencyGraph("example.xlsx")
graph.build_dependency_graph()
graph.print_graph()
print("Dependencies of Sheet1!A1:", graph.get_dependencies("Sheet1!A1"))
print("Dependents of Sheet1!A1:", graph.get_dependents("Sheet1!A1"))
