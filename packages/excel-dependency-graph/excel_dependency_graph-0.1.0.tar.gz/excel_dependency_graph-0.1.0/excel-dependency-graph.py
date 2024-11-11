import openpyxl
import re
import networkx as nx
import matplotlib.pyplot as plt
from typing import List, Dict, Set

class ExcelDependencyGraph:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=False)
        self.graph = nx.DiGraph()
        # Pattern matches single cells (e.g., A1), ranges (e.g., A1:B2), and supports absolute refs
        self.formula_pattern = r'(\b[A-Z]+\$?\d+\b)'

    def parse_worksheet(self, sheet_name: str) -> None:
        """Parse a worksheet and add cells and dependencies to the graph."""
        sheet = self.wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cell_coord = f"{sheet_name}!{cell.coordinate}"
                    formula = cell.value
                    dependencies = self._extract_dependencies(formula, sheet_name)
                    
                    # Add each dependency as an edge in the graph
                    for dep in dependencies:
                        self.graph.add_edge(dep, cell_coord)

    def _extract_dependencies(self, formula: str, sheet_name: str) -> Set[str]:
        """Extract cell dependencies from a formula using regex and return them as full sheet-cell refs."""
        matches = re.findall(self.formula_pattern, formula)
        dependencies = set()
        for match in matches:
            # Ensure each dependency is prefixed with the sheet name
            if '!' not in match:
                match = f"{sheet_name}!{match}"
            dependencies.add(match)
        return dependencies

    def build_dependency_graph(self) -> None:
        """Build the dependency graph by parsing each sheet in the workbook."""
        for sheet_name in self.wb.sheetnames:
            self.parse_worksheet(sheet_name)

    def visualize_graph(self) -> None:
        """Visualize the dependency graph."""
        plt.figure(figsize=(100,100))
        pos = nx.shell_layout(self.graph)
        nx.draw_networkx(self.graph, pos, with_labels=True, node_size = 5000, node_color="skyblue", font_size=10, arrowsize = 15,  arrows=True)
        plt.title("Excel Cell Dependency Graph")
        plt.show()

    def get_dependencies(self, cell: str) -> List[str]:
        """Return a list of cells that the given cell depends on."""
        if cell in self.graph:
            return list(self.graph.predecessors(cell))
        return []

    def get_dependents(self, cell: str) -> List[str]:
        """Return a list of cells that depend on the given cell."""
        if cell in self.graph:
            return list(self.graph.successors(cell))
        return []

    def save_graph(self, output_path: str) -> None:
        """Save the dependency graph to a file in GML format."""
        nx.write_gml(self.graph, output_path)

