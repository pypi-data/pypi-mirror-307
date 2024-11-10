"""
This script extracts formulas from an Excel file and builds a dependency graph.
"""

from typing import List, Dict
from openpyxl import load_workbook
import networkx as nx
import re
import sys
from .excel_parser import extract_references
import logging

logger = logging.getLogger(__name__)


# Dictionary that stores the unique functions used in the formulas
# The key will be the function name and the value will be the number of times it was used
functions_dict: Dict[str, int] = {}


def build_graph_and_stats(
    file_path: str,
    as_directed: bool = False,
) -> tuple[nx.DiGraph, Dict[str, int]]:
    """
    Extract formulas from an Excel file and build a dependency graph.
    """
    try:
        wb = load_workbook(file_path, data_only=False, read_only=True)
    except Exception as e:
        logger.error(f"Error loading workbook: {e}")
        sys.exit(1)

    graph = nx.DiGraph()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        logger.debug(f"========== Analyzing sheet: {sheet_name} ==========")
        sanitized_sheet_name = sanitize_sheetname(sheet_name)
        process_sheet(ws, sanitized_sheet_name, graph)

    if not as_directed:
        # Convert the graph to an undirected graph
        graph = graph.to_undirected()
    else:
        logger.info("Preserving the graph as a directed graph.")

    # remove selfloops
    graph.remove_edges_from(nx.selfloop_edges(graph))

    graph.remove_nodes_from(list(nx.isolates(graph)))

    return graph, functions_dict


def sanitize_sheetname(sheetname: str) -> str:
    """
    Remove any special characters from the sheet name.
    """
    return sheetname.replace("'", "")


def sanitize_nodename(nodename: str) -> str:
    """
    Remove any special characters from the node name.
    """
    return nodename.replace("'", "")


def sanitize_range(rangestring: str) -> str:
    """
    Remove any special characters from the range.
    """
    if "!" in rangestring:
        sheet, range_ = rangestring.split("!")
        sheet = sheet.replace("'", "")
        return f"{sheet}!{range_}"
    return rangestring


def stat_functions(cellvalue: str) -> None:
    """
    Extract the functions used in the formula and store them in a dictionary.
    This will be used to print the most used functions in the formulas.
    """
    cellfuncs = re.findall(r"[A-Z]+\(", cellvalue)
    logger.debug(f"  Functions used: {functions_dict}")
    for function in cellfuncs:
        function = function[:-1]  # Remove the "(" from the function name
        functions_dict[function] = functions_dict.get(function, 0) + 1


def add_node(graph: nx.DiGraph, node: str, sheet: str) -> None:
    """
    Add a node to the graph with the specified sheet name.
    """
    logger.debug(f"Adding node: {node} in sheet: {sheet}")
    sheet = sanitize_sheetname(sheet)
    node = sanitize_nodename(node)
    graph.add_node(node, sheet=sheet)


def process_sheet(ws, sheet_name: str, graph: nx.DiGraph) -> None:
    """
    Process a sheet and add references to the graph.
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                process_formula_cell(cell, sheet_name, graph)


def process_formula_cell(cell, sheet_name: str, graph: nx.DiGraph) -> None:
    """
    Process a cell containing a formula.
    """
    stat_functions(cell.value)
    cell_reference = f"{sheet_name}!{cell.coordinate}"
    logger.debug(f"Formula in {cell_reference}: {cell.value}")
    add_node(graph, cell_reference, sheet_name)

    direct_references, range_references, range_dependencies = extract_references(
        cell.value
    )
    add_references_to_graph(direct_references, cell_reference, sheet_name, graph)
    add_ranges_to_graph(range_references, cell_reference, sheet_name, graph)
    add_range_dependencies_to_graph(range_dependencies, sheet_name, graph)


def add_references_to_graph(
    references: List[str], current_cell: str, sheet_name: str, graph: nx.DiGraph
) -> None:
    """
    Add direct cell references to the graph.
    """
    for cell_reference in references:
        cell_reference = format_reference(cell_reference, sheet_name)
        logger.debug(f"  Cell: {cell_reference}")
        add_node(graph, cell_reference, sheet_name)
        graph.add_edge(current_cell, cell_reference)


def add_ranges_to_graph(
    ranges: List[str], current_cell: str, sheet_name: str, graph: nx.DiGraph
) -> None:
    """
    Add range references to the graph.
    """
    for range_reference in ranges:
        range_sheet_name = get_range_sheet_name(range_reference, sheet_name)
        range_reference = format_reference(range_reference, sheet_name)
        logger.debug(f"  Range: {range_reference}")
        add_node(graph, range_reference, range_sheet_name)
        graph.add_edge(current_cell, range_reference)


def add_range_dependencies_to_graph(
    range_dependencies: Dict[str, str], sheet_name: str, graph: nx.DiGraph
) -> None:
    """
    Add dependencies between ranges and cells.
    """
    for cell_reference, range_reference in range_dependencies.items():
        range_reference = format_reference(range_reference, sheet_name)
        cell_reference = format_reference(cell_reference, sheet_name)
        range_sheet_name = range_reference.split("!")[0]
        cell_sheet_name = cell_reference.split("!")[0]

        add_node(graph, cell_reference, cell_sheet_name)
        add_node(graph, range_reference, range_sheet_name)
        graph.add_edge(range_reference, cell_reference)


def format_reference(reference: str, sheet_name: str) -> str:
    """
    Format a cell or range reference to include the sheet name if not already present.
    """
    return (
        f"{sheet_name}!{reference}"
        if "!" not in reference
        else reference.replace("'", "")
    )


def get_range_sheet_name(range_reference: str, sheet_name: str) -> str:
    """
    Get the sheet name for a range reference.
    """
    return sheet_name if "!" not in range_reference else range_reference.split("!")[0]
