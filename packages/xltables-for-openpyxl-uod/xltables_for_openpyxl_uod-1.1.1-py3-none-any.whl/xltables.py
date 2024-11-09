from typing import Generator, Callable, Dict, Any, Tuple, List, Type, overload

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

class XLTable:
    """
    Utility wrapper class for openpyxl's Table class, adding some functionality or making some functionality more easily
    accessible.
    """

    file_path: str | None
    """The filepath of the workbook (.xlsx file) the table this represents is in."""

    source_workbook: Workbook
    """The workbook (.xlsx file) the table this represents is in."""

    source_worksheet: Worksheet
    """The worksheet the table this represents is in."""

    source_table: Table
    """The table this instance of XLTable extends.."""

    @property
    def ref(self) -> str:
        """
        The table range reference string. e.g. "A2:C20".
        :return: The table range reference string. e.g. "A2:C20".
        """
        return self.source_table.ref

    min_column_letters: str = ""
    """
    The name of the earliest column in the table. This is the column number as it appears in Excel- e.g. "AA" would be
    the 27th column.
    """

    max_column_letters: str = ""
    """
    The name of the last column in the table. This is the column number as it appears in Excel- e.g. "AA" would be
    the 27th column.
    """

    @property
    def column_names(self) -> list[str]:
        """
        The names of the columns in this table.
        :return: The names of the columns in this table. The is, the text appearing in the header of each column, in
                 order of column number.
        """

        result: list[str] = []
        header_row_number = self.max_header_row_number

        for i in range(self.min_column_number_in_sheet, self.max_column_number_in_sheet + 1):
            result.append(str(self.source_worksheet.cell(header_row_number, i).value))

        return result

    min_column_number_in_sheet: int = -1
    """The column number in the sheet of the earliest column in the table."""

    max_column_number_in_sheet: int = -1
    """The column number in the sheet of the last column in the table."""

    @property
    def column_count(self) -> int:
        """
        The number of columns in this table.
        :return: The number of columns in this table.
        """

        return self.max_column_number_in_sheet - self.min_column_number_in_sheet + 1

    @property
    def last_column(self) -> tuple[Cell, list[Cell]]:
        """
        The last column in the table.
        :return: The contents of the last column in the table as a list, along with the cell that holds the column's
                 name.
        """

        return self.get_column(self.column_count - 1)

    def column_iterator(self) -> Generator[tuple[str, list[Cell]], None, None]:
        """
        An iterator over the names then the contents of the columns in this table.
        :return: An iterator that produces each column name in this table, paired with a list of cells where the cells
                 in order from the cell at the top of the column to the cell at the bottom.
        """

        for col_no in range(self.min_column_number_in_sheet, self.max_column_number_in_sheet + 1):
            column_header_cell: Cell       = self.source_worksheet.cell(self.max_header_row_number, col_no)
            column_name:        str        = str(column_header_cell.value)
            column_cells:       list[Cell] = []

            for i in range(self.min_row_number_in_sheet, self.max_row_number_in_sheet + 1):
                column_cells.append(self.source_worksheet.cell(i, col_no))

            yield (column_name, column_cells)

    min_row_number_in_sheet: int = -1
    """The row number in the sheet of the earliest row in the table."""

    max_row_number_in_sheet: int = -1
    """The row number in the sheet of the last row in the table."""

    @property
    def row_count(self) -> int:
        """
        The number of rows in this table.
        :return: The number of rows in this table.
        """

        return self.max_row_number_in_sheet - self.min_row_number_in_sheet + 1

    @property
    def bottom_row(self) -> dict[str, Any]:
        """
        The last row in the table.
        :return: The contents of the last row in the table as a dictionary, where the keys are the column names (i.e.
                 the text that appears in the column's header cell) and the values are the contents of the cells in the
                 corresponding columns of the row.
        """

        return self.get_row(self.row_count - 1)

    @property
    def row_iterator(self) -> Generator[dict[str, Cell], None, None]:
        """
        An iterator over the rows of the table, excluding header rows.
        :return: An iterator over the rows of the table, as dictionaries of cells where the key is the column header.
                 Note that this assumes the table has at least one header row, and assumes that the final header row is
                 the one containing the column titles.
        """

        column_names = self.column_names

        for row in self.row_as_tuple_iterator:
            cells: dict[str, Cell] = {}

            for i in range(len(column_names)):
                cells[column_names[i]] = row[i]

            yield cells

    @property
    def row_as_tuple_iterator(self) -> Generator[tuple[Cell, ...], None, None]:
        """
        Provides an iterator over the rows of the table, excluding any header rows.
        :return: An iterator over the rows of the table, as tuples of cells as provided by `Worksheet.iter_rows`.
        """

        return self.source_worksheet.iter_rows(self.min_row_number_in_sheet,
                                               self.max_row_number_in_sheet,
                                               self.min_column_number_in_sheet,
                                               self.max_column_number_in_sheet)

    @property
    def min_header_row(self) -> int:
        """
        The row number in the sheet of the first header row of this table.
        :return: The row number in the sheet of the first header row of this table.
        """

        return self.min_row_number_in_sheet - self.header_row_count

    @property
    def max_header_row_number(self) -> int:
        """
        The row number in the sheet of the last header row of this table.
        :return: The row number in the sheet of the last header row of this table.
        """

        return self.min_row_number_in_sheet - 1

    @property
    def header_row_count(self) -> int:
        """
        The number of header rows of this table.
        :return: The number of header rows of this table.
        """

        return self.source_table.headerRowCount

    def __init__(self, file_path: str, workbook: Workbook, worksheet: Worksheet, table: Table):

        self.filepath         = file_path
        self.source_workbook  = workbook
        self.source_worksheet = worksheet
        self.source_table     = table
        splits: list[str]     = table.ref.split(":")

        top_left:     str = splits[0]
        bottom_right: str = splits[1]

        for i in range(len(top_left)):
            c: str = top_left[i]

            if(c.isdigit()):
                self.min_column_letters = top_left[:i]
                self.min_column_number_in_sheet = self.__convert_alphabetic_number_to_int(self.min_column_letters)
                self.min_row_number_in_sheet = int(top_left[i:]) + table.headerRowCount
                break

        for i in range(len(bottom_right)):
            c: str = bottom_right[i]

            if(c.isdigit()):
                self.max_column_letters = bottom_right[:i]
                self.max_column_number_in_sheet = self.__convert_alphabetic_number_to_int(self.max_column_letters)
                self.max_row_number_in_sheet = int(bottom_right[i:])
                break

    @staticmethod
    def load_from_file(file_path: str, sheet_name: str, table_name: str):
        """
        Creates a new XLTable representing a table in an Excel file, given a file path, sheet name, and table name.
        :param file_path: The file path of the Excel file to load the table from.
        :param sheet_name: The name of the sheet that contains the table.
        :param table_name: The name of the table.
        :return: A new XLTable object representing a table in the Excel file at the given path, in the sheet of the
                 given name, with the given table name.
        :raises InvalidFileException: If the specified file is not a workbook file readable by OpenPyXL.
        :raises ValueError: If the sheet name or table name do not lead to a valid table.
        """

        wb: Workbook = openpyxl.load_workbook(file_path, data_only = True)
        ws: Worksheet = wb[sheet_name]

        if(not isinstance(ws, Worksheet)):
            raise ValueError(f"No sheet in the specified workbook with the name \"{sheet_name}\".")

        opxl_table: Table | None = ws.tables.get(table_name)

        if(opxl_table is None):
            raise ValueError(f"No table in the specified sheet with the name \"{table_name}\".")

        return XLTable(file_path, wb, ws, opxl_table)

    def get_row(self, row_number: int) -> dict[str, Cell]:
        """
        Gets the nth row in the table. Note that the row number here is zero-indexed, and begins at the first data row.
        (under the header row)
        :param row_number: The zero-indexed number of the row to get.
        :return: The contents of the row at the given row number in the table as a dictionary, where the keys are the
                 column names (i.e. the text that appears in the column's header cell) and the values are the contents
                 of the cells in the corresponding columns of the row.
        """

        column_names:        list[str]       = self.column_names
        row_number_in_sheet: int             = self.min_row_number_in_sheet + row_number
        result:              dict[str, Cell] = {}

        for i in range(self.column_count):
            column_number_in_sheet: int = self.min_column_number_in_sheet + i
            result[column_names[i]]     = self.source_worksheet.cell(row_number_in_sheet, column_number_in_sheet)

        return result

    def get_column(self, column_name_or_number: str | int) -> tuple[Cell, list[Cell]]:
        """
        Gets the nth column in the table, or the column in the table with the matching name. Note that the column number
        here is zero-indexed, and begins at the first column of the table. (not the sheet)
        :param column_name_or_number: The name of the column to get.
        :return: A tuple, containing the cell with the name of the column to get, and a list of the cells in the same
                 order they appear in the table, from top to bottom.
        """

        if(column_name_or_number is str):
            column_name:       str = column_name_or_number
            header_row_number: int = self.max_header_row_number

            for i in range(self.column_count):
                cell: Cell = self.source_worksheet.cell(header_row_number, self.min_column_number_in_sheet + i)

                if(str(cell.value) == column_name):
                    return self.get_column(i)

            raise ValueError(f"No column in the table with the name \"{column_name}\"")

        if(column_name_or_number is int):
            column_number: int = column_name_or_number

            header_cell: Cell = self.source_worksheet.cell(self.max_header_row_number,
                                                          self.min_column_number_in_sheet + column_number)


            # column_name: str = str(header_cell.value)
            cells: list[Cell] = []

            for i in range(self.row_count):
                cells.append(self.source_worksheet.cell(self.min_row_number_in_sheet + i,
                                                        self.min_column_number_in_sheet + column_number))

            return (header_cell, cells)

    def get_column_number(self, column_name: str) -> int | None:
        header_row_number: int = self.max_header_row_number

        for i in range(self.column_count):
            cell: Cell = self.source_worksheet.cell(header_row_number, self.min_column_number_in_sheet + i)

            if(str(cell.value) == column_name):
                return i

    def get_column_name(self, column_number: int) -> str:
        column_number_in_sheet: int  = self.min_column_number_in_sheet + column_number
        header_row_number:      int  = self.max_header_row_number
        column_name_cell:       Cell = self.source_worksheet.cell(header_row_number, column_number_in_sheet)
        return str(column_name_cell.value)

    def has_column(self, column_name: str) -> bool:
        """
        Gets whether a column exists in this table with a given name.
        :param column_name: The name of the column to look for.
        :return: True if any column has the given name - that is, any column's header cell matches the given text.
                 Otherwise, false.
        """

        header_row_number = self.max_header_row_number

        for i in range(self.column_count):
            column_number: int = self.min_column_number_in_sheet + i

            if(str(self.source_worksheet.cell(header_row_number, column_number).value) == column_name):
                return True

    def add_row(self, values: dict[str, Any] | list[Any] | None = None) -> None:
        """
        Adds a row to the table this represents. This extends the table down one row in the sheet it's in. The cells are
        populated with values from the passed dictionary where the key is the same as the column header for that cell's
        column.
        :raises ValueError: If any key of the dictionary does not correspond to a value in the table.
        :param values: A dictionary of values to place in the new row, in cells in column with headers with column names
                       that match the key corresponding to that value.
        """

        self.max_row_number_in_sheet += 1
        self.source_table.ref = (f"{self.min_column_letters}{self.min_row_number_in_sheet - self.header_row_count}"
                                 f":{self.max_column_letters}{self.max_row_number_in_sheet}")

        if(values is None):
            return

        if(XLTable.__value_is_or_inherits_from(values, list[Any])):
            values: list[Any]

            for i in range(len(values)):
                self.source_worksheet.cell(self.max_row_number_in_sheet, self.min_column_number_in_sheet + i).value \
                    = values[i]

        if(XLTable.__value_is_or_inherits_from(values, dict[str, Any])):
            values: dict[str, Any]
            new_row: dict[str, Cell] = self.bottom_row

            for k, v in values.items():
                cell: Cell = new_row.get(k)

                if(cell is None):
                    raise ValueError(f"The column \"{k}\" does not exist in this table.")

                cell.value = v

    def delete_row(self, row_number: int) -> None:
        """
        Deletes a row from the table at the given row number. This will shift all rows below up by one to fill the gap,
        and result in the table being one row shorter.
        :param row_number: The number of the row to delete.
        """

        for i in range((row_number + 1), self.row_count):
            this_row = self.get_row(i)
            prev_row = self.get_row(i - 1)

            for k, v in this_row.items():
                prev_row[k].value = v.value

        for v in self.bottom_row.values():
            v.value = None

        self.max_row_number_in_sheet -= 1
        self.source_table.ref = (f"{self.min_column_letters}{self.min_row_number_in_sheet - self.header_row_count}"
                                 f":{self.max_column_letters}{self.max_row_number_in_sheet}")

    def delete_rows(self, row_numbers: list[int]) -> None:
        """
        Deletes the rows from the table at the row numbers in the given list. This will shift all rows below the rows
        deletes to fill the gap, and result in the table being as many rows shorter as there were rows removed.
        :param row_numbers: The numbers of the rows to remove.
        """

        if(len(row_numbers) == 0):
            return

        row_numbers                   = sorted(row_numbers)
        first_changed_row_number: int = row_numbers[0]
        adjustment:               int = 1

        for i in range((first_changed_row_number + 1), self.row_count):
            if(i in row_numbers):
                adjustment += 1
                continue

            this_row = self.get_row(i)
            row_being_overwritten = self.get_row(i - adjustment)

            for k, v in this_row.items():
                row_being_overwritten[k].value = v.value

        row_count = self.row_count

        for i in range(row_count - adjustment, row_count):
            for v in self.get_row(i).values():
                v.value = None

        self.max_row_number_in_sheet -= adjustment
        self.source_table.ref = (f"{self.min_column_letters}{self.min_row_number_in_sheet - self.header_row_count}"
                                 f":{self.max_column_letters}{self.max_row_number_in_sheet}")

    def delete_rows_where(self, test: Callable[[dict[str, Cell]], bool]) -> None:
        """
        Deletes the rows from the table that match the given predicate. This will shift all rows below the rows deleted
        to fill the gap, and result in the table being as many rows shorter as there were rows removed.
        :param test: The predicate to check rows against to see if they should be removed.
        """

        adjustment: int = 0
        row_count: int = self.row_count

        for i in range(0, self.row_count):
            this_row: dict[str, Cell] = self.get_row(i)

            if(test(this_row)):
                adjustment += 1
                continue

            if(adjustment == 0):
                continue

            row_being_overwritten = self.get_row(i - adjustment)

            for k, v in this_row.items():
                row_being_overwritten[k].value = v.value

        for i in range((row_count - adjustment), row_count):
            for v in self.get_row(i).values():
                v.value = None

        self.max_row_number_in_sheet -= adjustment
        self.source_table.ref = (f"{self.min_column_letters}{self.min_row_number_in_sheet - self.header_row_count}"
                                 f":{self.max_column_letters}{self.max_row_number_in_sheet}")

    def add_column(self, column_name: str, values: list[Any] | None = None) -> None:
        """
        Adds a column to the table this represents. This extends the table one column to the right in the sheet it's in.
        :param column_name: The name of the column. This will appear in the column header.
        :param values: The values to populate the column with.
        """

        new_column_number: int = self.max_column_number_in_sheet + 1
        self.source_worksheet.cell(self.max_header_row_number, new_column_number).value = column_name

        if(values is not None):
            for i in range(len(values)):
                self.source_worksheet.cell(self.min_row_number_in_sheet + i, new_column_number).value = values[i]

        self.max_column_number_in_sheet += 1
        self.max_column_letters = XLTable.__convert_int_to_alphabetic_number(self.max_column_number_in_sheet)
        self.source_table.ref = (f"{self.min_column_letters}{self.min_row_number_in_sheet - self.header_row_count}"
                                 f":{self.max_column_letters}{self.max_row_number_in_sheet}")

    def delete_column(self, column_name_or_number: str | int) -> None:
        """
        Deletes a column from the table with the given column number (from the start of the table) or name. This will
        shift all columns left by one to fill the gap, and result in the table being one column shorter.
        :param column_name_or_number:
        :return:
        """

        if(column_name_or_number is str):
            self.delete_column(self.get_column_number(column_name_or_number))
            return

        if(column_name_or_number is int):
            column_number: int = column_name_or_number

            for column_number in range((column_number + 1), self.column_count):
                this_header_cell, this_cells = self.get_column(column_number)
                prev_header_cell, prev_cells = self.get_column(column_number - 1)

                prev_header_cell.value = this_header_cell.value

                for row_number in range(self.row_count):
                    prev_cells[row_number] = this_cells[row_number]

            for v in self.last_column[1]:
                v.value = None

            self.max_column_number_in_sheet -= 1
            self.source_table.ref = None # TO DO: Write.

    def save(self, filepath: str | None = None):
        """
        Saves the table, committing any changes made to the file.

        Note that this saves the entire .xlsx file the table is in - if any changes have been made elsewhere, this will
        save them as well.
        :param filepath: The filepath to save at. By default, (that is, if None) this method saves to the same filepath
                         the table was loaded from.
        """

        self.source_workbook.save((filepath) if (filepath is not None) else (self.file_path))

    @staticmethod
    def __convert_int_to_alphabetic_number(to_convert: int) -> str:
        """
        Converts an integer to a string representation of it in an alphabetic format. (i.e. the format that Excel uses
        for columns) (e.g. Z = 26, AA = 27)
        :param to_convert: The number to convert.
        :return: The string representation of the given number, in an alphabetic format. That is, base-26 where A = 1,
                 B = 2, C = 3, etc.
        """

        result_chars: list[str] = []

        while (to_convert > 0):
            to_convert -= 1
            result_chars.append(chr(to_convert % 26 + ord("A")))
            to_convert //= 26

        result_chars.reverse()
        return "".join(result_chars)

    @staticmethod
    def __convert_alphabetic_number_to_int(to_convert: str) -> int:
        """
        Converts a number in an alphabetic format (i.e. the format that Excel uses for columns) (e.g. Z = 26, AA = 27)
        to an integer.
        :param to_convert: A string containing an alphabetic representation of a number.
        :return: The integer representation of the passed alphabetic number.
        """

        i: int = 0
        l: int = len(to_convert)

        result: int = 0

        while (i < l):
            c: str = to_convert[l - i - 1]
            result += (ord(c) - 64) * (26 ** i)

            i += 1

        return result

    @staticmethod
    def __value_is_or_inherits_from(value: Any, type_that_value_may_be: Type[Any]) -> bool:
        """
        Tests a value to see if it's an instance of a given type. That is, if its type is the same or a subtype of the
        given type.
        :param value: The value to check the type of.
        :param type_that_value_may_be: The type that the value may be.
        :return: True if the given value's type is the same or a subtype of the given type. Otherwise, false.
        """

        if(value is type_that_value_may_be):
            return True

        return XLTable.__type_is_or_inherits_from(type(value), type_that_value_may_be)

    @staticmethod
    def __type_is_or_inherits_from(type_to_check: Type[Any], type_that_other_may_be: Type[Any]) -> bool:
        """
        Tests a type to see if it's an the same as, or a subtype of, of another type.
        :param type_to_check: The type to check to see if it's the same as, or a subtype of, the other.
        :param type_that_other_may_be: The type to check the other type against.
        :return: True if the first type is the same as, or a subtype of, the other. Otherwise, false.
        """

        if(type_to_check == type_that_other_may_be):
            return True

        for base_type in type_to_check.__bases__:
            if(XLTable.__type_is_or_inherits_from(base_type, type_that_other_may_be)):
                return True

        return False
