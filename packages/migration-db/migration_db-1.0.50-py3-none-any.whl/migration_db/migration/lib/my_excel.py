"""
@Author: xiaodong.li
@Date: 2020-04-10 13:10:45
@LastEditTime : 2020-04-22 13:56:05
@LastEditors  : Please set LastEditors
@Description: Translate excel.
@FilePath: \\trans_excel\\trans.py
"""

from openpyxl import load_workbook


class ImportDto:

    def __init__(self, columns_len):
        import_var = self.__dict__
        for i in range(columns_len):
            import_var['var' + str(i)] = i


class MyExcel:
    """excel."""

    def __init__(self, _path, sheet_name='Sheet1'):
        self.wb = load_workbook(filename=_path)
        self.ws = self.wb[sheet_name]
        self.max_row = self.ws.max_row
        self.max_column = self.ws.max_column
        self.start_row = None
        self.end_row = None

    def row_generator(self, columns=None, start_row=2, end_row=None):
        max_row = self.max_row
        if end_row is None or (end_row > max_row):
            end_row = max_row
        elif end_row < 2:
            end_row = 2
        elif isinstance(end_row, str) and end_row.lower() == 'more':
            end_row = max_row + 1
        if (start_row < 2) or (start_row > max_row):
            start_row = 2
        self.start_row = start_row
        self.end_row = end_row
        return self._row_generator(columns)

    def _row_generator(self, columns):
        columns_len = len(columns)
        for i in range(columns_len):
            exec(f'index{i} = -1')
        for i in range(columns_len):
            for column in range(self.max_column):
                text = self._revise_val(self.ws[1][column].value)
                if text == columns[i]:
                    exec(f'index{i} = column')
                    break
        for i in range(columns_len):
            if eval(f"index{i} == -1"):
                raise Exception(
                    "Please check the excel: "
                    "incoming parameters: columns do not match the 'column names' in the document.")
        for row in range(self.start_row, self.end_row + 1):
            import_dto = ImportDto(columns_len)
            for i in range(columns_len):
                exec(f"import_dto.var{i} = self._revise_val(self.ws[row][index{i}].value)")
            yield import_dto

    @staticmethod
    def _revise_val(val):
        if isinstance(val, str):
            return val.strip()
        return val
