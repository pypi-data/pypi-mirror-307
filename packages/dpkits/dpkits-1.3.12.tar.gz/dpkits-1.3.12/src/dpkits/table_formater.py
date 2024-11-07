import time
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.styles.fills import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import workbook, worksheet
from colorama import Fore
from .logging import Logging
import multiprocessing
import xlwings as xw
from xlwings import sheets




class TableFormatter(Logging):

    def __init__(self, xlsx_name):

        super().__init__()
        self.xlsx_name = xlsx_name.rsplit('/', 1)[-1] if '/' in xlsx_name else xlsx_name

        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        self.dot_border = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='dotted'))

        self.dot_thin_right_border = Border(left=Side(style='dotted'), right=Side(style='thin'), top=Side(style='dotted'), bottom=Side(style='dotted'))
        self.dot_thin_bot_border = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='thin'))
        self.dot_right_thin_right_border = Border(left=Side(style='dotted'), right=Side(style='thin'), top=Side(style='dotted'), bottom=Side(style='thin'))

        self.medium_left_1_border = Border(left=Side(style='medium'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='dotted'))
        self.medium_left_2_border = Border(left=Side(style='medium'), right=Side(style='dotted'), top=Side(style='dotted'), bottom=Side(style='thin'))
        self.medium_left_3_border = Border(left=Side(style='medium'), right=Side(style='dotted'), top=Side(style='thin'), bottom=Side(style='dotted'))
        self.medium_left_4_border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))



    @staticmethod
    def format_content_sheet(ws_content, ws_dtbl, str_cell_content_addr):

        idx = ws_content.max_row + 1

        ws_dtbl[str_cell_content_addr].value = 'Content'
        ws_dtbl[str_cell_content_addr].hyperlink = f"#'Content'!B{idx}"
        ws_dtbl[str_cell_content_addr].font = Font(color='0000FF')

        ws_content[f'A{idx}'].value = ws_content.max_row
        ws_content[f'B{idx}'].value = ws_dtbl.title
        ws_content[f'B{idx}'].hyperlink = f"#'{ws_dtbl.title}'!{str_cell_content_addr}"
        ws_content[f'B{idx}'].font = Font(color='0000FF')



    def format_side_axis(self, ws, is_matrix_table, start_side_row, num_format, step, lst_sub_header_col, is_tbl_sig=False):

        lst_sub_side_col = list()

        start_row = start_side_row

        for irow in range(start_row, ws.max_row + 1):

            cur1_cell = ws.cell(irow, 1)

            cur1_cell.border = self.thin_border
            cur1_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cur1_cell.font = Font(bold=True, color='FFFFFF')
            cur1_cell.fill = PatternFill(patternType='solid', fgColor='538DD5')

            cur2_cell = ws.cell(irow, 2)
            cur2_cell.border = self.thin_border
            cur2_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cur2_cell.font = Font(bold=True)
            cur2_cell.fill = PatternFill(patternType='solid', fgColor='D9E1F2')

            if cur2_cell.value != ws.cell(irow + 1, 2).value or cur1_cell.value != ws.cell(irow + 1, 1).value:

                if '_Group' in str(ws.cell(irow + 1, 2).value) or '_Mean' in str(ws.cell(irow + 1, 2).value):
                    continue

                end_row = irow
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                start_row = irow + 1
                lst_sub_side_col.append(irow)

        if is_matrix_table:
            print(f"Formatting {ws.title} - Brand label")

            for icol in range(5, ws.max_column):
                for row in range(5, 7):
                    cur_cell = ws.cell(row, icol)
                    cur_cell.border = self.thin_border
                    cur_cell.font = Font(bold=True)

                    if row == 5:
                        cur_cell.alignment = Alignment(horizontal='center', vertical='center')
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='D8E4BC')
                    elif row == 6:
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='DAEEF3')

        start_row = start_side_row + 1 if is_matrix_table else start_side_row

        for irow in range(start_row, ws.max_row):

            if irow % 500 == 0:
                print(f"Formatting {ws.title} - row {irow}")

            if 'FT' in ws.cell(irow, 3).value:
                ws.row_dimensions[irow].hidden = True
            else:
                for icol in range(5, ws.max_column):

                    cur_cell = ws.cell(irow, icol)

                    cur_cell.border = self.dot_border

                    if icol in lst_sub_header_col:
                        cur_cell.border = self.dot_thin_right_border

                    if irow in lst_sub_side_col:
                        cur_cell.border = self.dot_thin_bot_border

                    if icol in lst_sub_header_col and irow in lst_sub_side_col:
                        cur_cell.border = self.dot_right_thin_right_border


                    if is_tbl_sig and ws.cell(start_side_row - 1, icol).value == 'A':
                        cur_cell.border = self.medium_left_1_border

                        if irow in lst_sub_side_col:
                            cur_cell.border = self.medium_left_2_border

                        if icol in lst_sub_header_col and irow in lst_sub_side_col:
                            cur_cell.border = self.medium_left_3_border

                        for irow_hd in range(3, start_side_row):
                            ws.cell(irow_hd, icol).border = self.medium_left_4_border

                    if ws.cell(irow, 4).value == 'base':
                        cur_cell.font = Font(bold=True)
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='DAEEF3')

                    elif ws.cell(irow, 4).value == 'bes':
                        cur_cell.font = Font(bold=True)
                        cur_cell.number_format = '0.0'
                        cur_cell.fill = PatternFill(patternType='solid', fgColor='FDE9D9')

                        if icol == 6 or (icol > 6 and (icol - 6) % step == 0):
                            cur_cell.alignment = Alignment(horizontal='center', vertical='center')
                            ws.merge_cells(start_row=irow, start_column=icol, end_row=irow, end_column=icol + step - 1)

                    else:
                        if ws.cell(irow, 4).value in ['mean', 'std']:
                            ws.cell(irow, 5).font = Font(bold=True, color='0070C0')
                            cur_cell.font = Font(bold=True)
                            cur_cell.number_format = '0.00'

                        elif ws.cell(irow, 4).value in ['min', 'max']:
                            ws.cell(irow, 5).font = Font(bold=True, color='00B050')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif ws.cell(irow, 4).value in ['25%', '50%', '75%']:
                            ws.cell(irow, 5).font = Font(bold=True, color='963634')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif 'calculate' in ws.cell(irow, 4).value:
                            ws.cell(irow, 5).font = Font(bold=True, color='E26B0A')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif 'friedman_pval' in ws.cell(irow, 4).value:
                            ws.cell(irow, 5).font = Font(bold=True, color='E26B0A')
                            cur_cell.font = Font(italic=True)
                            cur_cell.number_format = '0.00'

                        elif ws.cell(irow, 3).value == 'GROUP':
                            cur_cell.font = Font(bold=True)

                            if num_format == 'pct':
                                cur_cell.number_format = '0'
                            elif num_format == 'pct_sign':
                                cur_cell.number_format = '0%'

                        elif int(ws.cell(irow, 4).value) > 90000:

                            cur_cell.font = Font(bold=True)

                            if num_format == 'pct':
                                cur_cell.number_format = '0'
                            elif num_format == 'pct_sign':
                                cur_cell.number_format = '0%'

                            cat_lbl = str(ws.cell(irow, 5).value)
                            net_cell = ws.cell(irow, icol)

                            # NET CODE UPDATE FOR MSN PROJECTS
                            if 'NET 0' in cat_lbl.upper():
                                # Net: #A02B93
                                net_cell.fill = PatternFill(patternType='solid', fgColor='A02B93')
                                net_cell.font = Font(bold=True, color='FFFFFF')

                            elif 'NET 1' in cat_lbl.upper():
                                # Sub-net: #F2CEEF
                                net_cell.fill = PatternFill(patternType='solid', fgColor='F2CEEF')

                            elif 'NET 2' in cat_lbl.upper():
                                # Sub-sub-net: #B5E6A2
                                net_cell.fill = PatternFill(patternType='solid', fgColor='B5E6A2')

                            else:
                                if '(NET)' in cat_lbl.upper():
                                    ws.cell(irow, 5).font = Font(bold=True, underline='double')
                                else:
                                    ws.cell(irow, 5).font = Font(bold=True, underline='single')

                        else:
                            if icol > 5:
                                if num_format == 'pct':
                                    cur_cell.number_format = '0'
                                elif num_format == 'pct_sign':
                                    cur_cell.number_format = '0%'

                        if is_tbl_sig and icol % 2 != 0 and icol > 5:
                            cur_cell.font = Font(bold=True, color='FF0000')



    def format_sig_table(self):

        wb = openpyxl.load_workbook(self.xlsx_name, data_only=True)

        ws_content = wb['Content']
        ws_content.column_dimensions['B'].width = 60

        for ws in wb.worksheets:

            if ws.title in ['Content']:
                continue

            print(f"Formatting {ws.title}")

            self.format_content_sheet(ws_content, ws, str_cell_content_addr='A3')

            lst_sub_header_col = list()

            last_header_row = 2

            while ws.cell(last_header_row, 4).value != 'base':
                last_header_row += 1

                if last_header_row >= 30:
                    last_header_row = -999
                    break

            last_header_row -= 1

            if not ws.cell(4, 2).value:
                is_sig_tbl = False
            else:
                is_sig_tbl = True if 'Dependent' in ws.cell(4, 2).value or 'Independent' in ws.cell(4, 2).value else False

            if not is_sig_tbl:
                ws.delete_rows(last_header_row - 1, 2)
                last_header_row = last_header_row - 2
            else:
                ws.delete_rows(last_header_row - 1, 1)
                last_header_row = last_header_row - 1

            ws.freeze_panes = f'F{last_header_row + 1}'

            if 'count' in ws.cell(3, 2).value:
                num_format = 'count'
            else:
                if '%' in ws.cell(3, 2).value:
                    num_format = 'pct_sign'
                else:
                    num_format = 'pct'

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['E'].width = 40

            if not is_sig_tbl:

                for icol in range(6, ws.max_column + 1).__reversed__():
                    if icol % 2 != 0:
                        # ws.delete_cols(icol, 1)
                        ws.delete_cols(icol)

            else:
                for icol in range(6, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(icol)].width = 7


            ws.column_dimensions['C'].hidden = True
            ws.column_dimensions['D'].hidden = True

            ws.row_dimensions[1].hidden = True
            ws.row_dimensions[2].hidden = True

            if is_sig_tbl and last_header_row > 1:
                ws.row_dimensions[last_header_row - 1].height = 30
            else:
                ws.row_dimensions[last_header_row].height = 30

            print(f"Formatting {ws.title} - Header")

            ws_max_col = ws.max_column

            dict_header_color = {
                0: {'fontColor': 'FFFFFF', 'fgColor': '203764'},
                1: {'fontColor': 'FFFFFF', 'fgColor': '305496'},
                2: {'fontColor': '000000', 'fgColor': '8EA9DB'},
                3: {'fontColor': '000000', 'fgColor': 'B4C6E7'},
                4: {'fontColor': '000000', 'fgColor': 'D9E1F2'},

                5: {'fontColor': '000000', 'fgColor': 'B5F1CD'},
                6: {'fontColor': '000000', 'fgColor': '57DF8E'},
                7: {'fontColor': '000000', 'fgColor': '27CF6B'},
                8: {'fontColor': 'FFFFFF', 'fgColor': '1E9E52'},
                9: {'fontColor': 'FFFFFF', 'fgColor': '156e3a'},

                # 5: {'fontColor': '000000', 'fgColor': 'A9D08E'},
                # 6: {'fontColor': '000000', 'fgColor': 'C6E0B4'},
                # 7: {'fontColor': '000000', 'fgColor': 'E2EFDA'},
                # 8: {'fontColor': '000000', 'fgColor': 'F4B084'},
                # 9: {'fontColor': '000000', 'fgColor': 'F8CBAD'},

                10: {'fontColor': '000000', 'fgColor': 'FCE4D6'},
                11: {'fontColor': '000000', 'fgColor': 'FCD5B4'},
                12: {'fontColor': '000000', 'fgColor': 'FDE9D9'},
            }

            for irow in range(3, last_header_row + 1):
                start_column = 6

                for icol in range(6, ws_max_col + 1):

                    cur_cell = ws.cell(irow, icol)

                    cur_cell.border = self.thin_border
                    cur_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cur_cell.font = Font(bold=True, size=12, color=dict_header_color[irow - 3]['fontColor'])
                    cur_cell.fill = PatternFill(patternType='solid', fgColor=dict_header_color[irow - 3]['fgColor'])

                    if irow <= 4:

                        if cur_cell.value != ws.cell(irow, icol + 1).value:
                            end_column = icol

                            if start_column != end_column and cur_cell.value is not None:

                                ws.merge_cells(start_row=irow, start_column=start_column, end_row=irow, end_column=end_column)
                                lst_sub_header_col.append(icol)

                            start_column = icol + 1

                    else:

                        if cur_cell.value != ws.cell(irow, icol + 1).value or ws.cell(irow - 1, icol + 1).value is not None:
                            end_column = icol

                            if start_column != end_column and cur_cell.value is not None:

                                ws.merge_cells(start_row=irow, start_column=start_column, end_row=irow, end_column=end_column)
                                lst_sub_header_col.append(icol)

                            start_column = icol + 1


            print(f"Formatting {ws.title} - Side axis")



            start_side_row = last_header_row + 1
            self.format_side_axis(ws=ws, is_matrix_table=False, start_side_row=start_side_row, num_format=num_format,
                                  step=1, lst_sub_header_col=lst_sub_header_col, is_tbl_sig=is_sig_tbl)


        # output_name = self.xlsx_name.replace('.xlsx', '_output.xlsx')
        output_name = self.xlsx_name

        print(f"Save wb as {output_name}")
        wb.save(output_name)
        wb.close()



    """
    --------------------------------------------------------------------------------------------------------------------
    ADD NEW-------------------------------------------------------------------------------------------------------------
    --------------------------------------------------------------------------------------------------------------------
    """



    def format_workbook(self):

        wb = xw.Book(self.xlsx_name)

        for ws in wb.sheets:
            self.format_worksheet(ws=ws)

        wb.save(self.xlsx_name.replace('.xlsx', '_Testing_Output.xlsx'))
        wb.close()


    def format_worksheet(self, ws: sheets):

        if ws.name == 'Content':
            self.format_ws_content(ws=ws)
        else:
            self.format_ws_table(ws=ws)




    def format_ws_content(self, ws: sheets):
        self.print(ws.name)

        """
        Script here
        """

        pass




    def format_ws_table(self, ws: sheets):

        self.print(ws.name)

        first_row = ws.range('A1').end('down').row
        first_column = ws.range('A2').end('right').column

        last_cell = ws.used_range.last_cell
        last_row = last_cell.row
        last_column = last_cell.column

        if ws.range('B4').value is None:
            lst_col_remove = list(range(first_column + 1, last_column + 1, 2))
            lst_col_remove.reverse()

            for icol in lst_col_remove:
                ws.range(f'{get_column_letter(icol)}:{get_column_letter(icol)}').api.Delete()

            ws.range(f'{first_row - 2}:{first_row - 1}').api.Delete()

            first_row = ws.range('A1').end('down').row
            first_column = ws.range('A2').end('right').column

            last_cell = ws.used_range.last_cell
            last_row = last_cell.row
            last_column = last_cell.column

        first_column_letter = get_column_letter(first_column)
        last_column_letter = get_column_letter(last_column)

        cells_range = ws.range(f'{first_column_letter}{first_row}:{last_column_letter}{last_row}')
        cells_range.api.Borders.Weight = 1  # dot


        header_range = ws.range(f'{first_column_letter}3:{last_column_letter}{first_row - 1}')
        header_range.api.Borders.Weight = 2  # thin line

        side_range = ws.range(f'A{first_row}:{get_column_letter(first_column - 1)}{last_row}')
        side_range.api.Borders.Weight = 2  # thin line


        # for i in range(7, 11):
        #     cells_range.api.Borders(i).Weight = 2

