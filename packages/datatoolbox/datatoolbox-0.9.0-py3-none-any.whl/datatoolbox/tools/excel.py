#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
This module contains all relevant tools regarding the use of excel


@author: ageiges
"""
import os
import re
import string
import pandas as pd
import tqdm
import platform
from pandas_indexing import accessors
from openpyxl import load_workbook
import numpy as np
from datatoolbox import config, mapp, core, io_tools, util
from shutil import copyfile
from copy import copy
#%% Defintions
OS = platform.system()  #'win32' , linux, #Darwin


REQUIRED_SETUP_FIELDS = [
    'filePath',
    'fileName',
    'sheetName',
    'timeIdxList',
    'spaceIdxList',
]

SP_ARG = 3
TI_ARG = 2
DT_ARG = 4

REG_EXCEL_RANGE = re.compile('^[A-Z]{1,3}[0-9]{1,3}:[A-Z]{1,3}[0-9]{1,3}$')
REG_EXCEL_ROW = re.compile('^[0-9]{1,2}$')
REG_EXCEL_COL = re.compile('^[A-Z]{1,2}$')

REG_FIND_ROWS = re.compile('^[A-Z]{1,3}([0-9]{1,3}):[A-Z]{1,3}([0-9]{1,3})$')
REG_FIND_COLS = re.compile('^([A-Z]{1,3})[0-9]{1,3}:([A-Z]{1,3})[0-9]{1,3}$')


#%% Private functions
def _isColRange(string):
    #    if REG_EXCEL_COL.match(string):
    #        return True
    match = REG_FIND_COLS.search(string)
    if match:
        if match.group(1) == match.group(2):
            return True

    return False


def _isRowRange(string):
    #    if REG_EXCEL_ROW.match(string):
    #        return True
    match = REG_FIND_ROWS.search(string)
    if match:
        if match.group(1) == match.group(2):
            return True

    return False


def _isRow(string):
    if REG_EXCEL_ROW.match(string):
        return True
    else:
        return False


def _isCol(string):
    if REG_EXCEL_COL.match(string):
        return True
    else:
        return False


#%%
def _iterTime(timeIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):
    if _isRowRange(timeIdxString):
        # colum setup
        for timeCell in wksheet[timeIdxString][0]:
            xlsCol = timeCell.col_idx
            timeIdx = timeCell.value
            # print(xlsCol, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isRow(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
            xlsCol = timeCell.col_idx
            timeIdx = timeCell.value
            # print(xlsRow, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    elif _isCol(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
            xlsRow = timeCell.row
            timeIdx = timeCell.value
            # print(xlsRow, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isColRange(timeIdxString):
        for timeCell in wksheet[timeIdxString]:
            #            print(timeCell[0].row)
            xlsRow = timeCell[0].row
            timeIdx = timeCell[0].value
            # print(xlsRow, timeIdx)
            #            print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed
        timeIdx = timeIdxString
        yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]


def _iterSpace(spaceIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):
    if _isRowRange(spaceIdxString):
        # colum setup
        for spaceCell in wksheet[spaceIdxString][0]:
            xlsCol = spaceCell.col_idx
            spaceIdx = spaceCell.value
            # print(xlsCol, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isRow(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsCol = spaceCell.col_idx
            spaceIdx = spaceCell.value
            # print(xlsRow, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isCol(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsRow = spaceCell.row
            spaceIdx = spaceCell.value
            # print(xlsRow, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isColRange(spaceIdxString):
        for spaceCell in wksheet[spaceIdxString]:
            xlsRow = spaceCell[0].row
            spaceIdx = spaceCell[0].value
            # print(xlsRow, spaceIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed string
        print('fixed')
        spaceIdx = spaceIdxString
        yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]


def _iterData(dataIdxString, wksheet, xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx):

    # print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])
    if _isRowRange(dataIdxString):
        # colum setup
        for dataCell in wksheet[dataIdxString][0]:
            xlsCol = dataCell.col_idx
            dataIdx = dataCell.value
            # print(xlsCol, timeIdx, dataIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isRow(dataIdxString):
        for dataCell in wksheet[dataIdxString]:
            xlsCol = dataCell.col_idx
            dataIdx = dataCell.value
            # print(xlsRow, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isCol(dataIdxString):
        #        print('_isCol')
        #        print([xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx])

        for i, dataCell in enumerate(wksheet[dataIdxString]):
            #            print(dataCell)
            #            print(dataCell.col_idx)
            #            xlsRow = dataCell.row
            xlsCol = dataCell.col_idx
            if i == 0:
                dataIdx = dataCell.value
            # print(xlsRow, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]

    elif _isColRange(dataIdxString):
        #        print('_isColRange')
        for dataCell in wksheet[dataIdxString]:
            # print(dataCell)
            xlsRow = dataCell[0].row
            if dataCell[0].value is not None:
                dataIdx = dataCell[0].value.replace('\ufeff', '')
            else:
                dataIdx = None
            # print(xlsRow, timeIdx)
            yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]
    else:
        # assume fixed
        dataIdx = dataIdxString
        yield [xlsRow, xlsCol, timeIdx, spaceIdx, dataIdx]


def _identifyCountry(string):
    """
    Used various methods to find best fitting country alpha3 iso code
    """

    import datatoolbox as dt

    # numeric ISO code
    try:
        numISO = float(string)
        mask = numISO == dt.mapp.countries.codes['numISO']
        if mask.any():
            return dt.mapp.countries.index[mask][0]
    except:
        pass
    if len(string) == 2:
        mask = string == dt.mapp.countries.codes['alpha2']
        if mask.any():
            return dt.mapp.countries.codes.index[mask][0]

    if len(string) == 3:

        if string.upper() in dt.mapp.countries.codes.index:
            return string.upper()

    try:
        coISO = dt.getCountryISO(string)
        return coISO
    except:
        print('not matching country found')
        return None


def _res2Excel(resFull, countryList):
    import datatoolbox as dt

    tableSet = dt.TableSet()

    yearRange = (pd.np.inf, -pd.np.inf)

    ID_list = list()
    for ID in resFull:
        table = dt.getTable(ID)
        if 'Gg' in table.meta['unit']:
            table = table.convert(table.meta['unit'].replace('Gg', 'Mt'))
        tableSet.add(table)
        ID_list.append(table.ID)

        minYears = min(yearRange[0], table.columns.min())
        maxYears = max(yearRange[1], table.columns.max())
        yearRange = (minYears, maxYears)

    for country in countryList:

        coISO = _identifyCountry(country)

        if coISO is None:
            coISO = country

        outDf = pd.DataFrame(
            columns=['Source', 'Entity', 'Category', 'Scenario', 'Unit']
            + list(range(yearRange[0], yearRange[1] + 1))
        )

        i = 0
        for ID in ID_list:
            if coISO in tableSet[ID].index:
                years = tableSet[ID].columns[~tableSet[ID].loc[coISO].isna()]
                outDf.loc[i, years] = tableSet[ID].loc[coISO, years]
                outDf.loc[i, ['Source', 'Entity', 'Category', 'Scenario', 'Unit']] = [
                    tableSet[ID].meta[x]
                    for x in ['source', 'entity', 'category', 'scenario', 'unit']
                ]
                i = i + 1

        outDf = outDf.loc[:, ~outDf.isnull().all(axis=0)]
        outDf.to_excel('extract_' + coISO + '.xlsx')


def _str2float(x):
    if isinstance(x, float) or isinstance(x, int) or x is None:
        return x
    if '%' in x:
        return float(x.replace('%', '')) * 100
    else:
        if x.startswith('#') or x == '':
            return np.nan
        #        try:
        print(x)
        return float(x)


#%% Functions
getAllSheetNames = io_tools.getAllSheetNames

def alphaCol2Num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num - 1


def colNameToNum(name):
    pow = 1
    colNum = 0
    for letter in name[::-1]:
        colNum += (int(letter, 36) - 9) * pow
        pow *= 26
    return colNum - 1


def compare_excel_files(file1, file2, eps=1e-6):
    """
    This function compares two excel files and creates a file visualising the differences

    """

    def report_diff(x):
        try:
            #            print(x)
            x = x.astype(float)
            #            print(x)
            return (
                x[1]
                if (abs(x[0] - x[1]) < eps) or pd.np.any(pd.isnull(x))
                else '{0:.2f} -> {1:.2f}'.format(*x)
            )
        except:
            return (
                x[1]
                if (x[0] == x[1]) or pd.np.any(pd.isnull(x))
                else '{} -> {}'.format(*x)
            )
            # print(x)

    xlFile1 = pd.ExcelFile(file1)
    sheetNameList1 = set(xlFile1.sheet_names)

    xlFile2 = pd.ExcelFile(file2)
    sheetNameList2 = set(xlFile2.sheet_names)

    writer = pd.ExcelWriter("diff_temp.xlsx")

    for sheetName in sheetNameList2.intersection(sheetNameList1):
        data1 = pd.read_excel(xlFile1, sheet_name=sheetName)
        data2 = pd.read_excel(xlFile2, sheet_name=sheetName)
        data1 = data1.replace(pd.np.nan, '').astype(str)
        data2 = data2.replace(pd.np.nan, '').astype(str)
        data1 = data1.apply(lambda x: x.str.strip())
        data2 = data2.apply(lambda x: x.str.strip())
        diff_panel = pd.Panel(dict(df1=data1, df2=data2))
        diff_output = diff_panel.apply(report_diff, axis=0)

        diff_output.to_excel(writer, sheet_name=sheetName)

        workbook = writer.book
        worksheet = writer.sheets[sheetName]
        highlight_fmt = workbook.add_format(
            {'font_color': '#FF0000', 'bg_color': '#B1B3B3'}
        )

        grey_fmt = workbook.add_format({'font_color': '#8d8f91'})

        worksheet.conditional_format(
            'A1:ZZ1000',
            {
                'type': 'text',
                'criteria': 'containing',
                'value': '->',
                'format': highlight_fmt,
            },
        )
        worksheet.conditional_format(
            'A1:ZZ1000',
            {
                'type': 'text',
                'criteria': 'not containing',
                'value': '->',
                'format': grey_fmt,
            },
        )

    writer.close()
    os.system('libreoffice ' + "diff_temp.xlsx")
    os.system('rm diff_temp.xlsx')


def excel_to_pandas_idx(index):
    """
    Conversion from Excel Row-Col sting to pandas index

    E.g. excel_to_pandas_idx('B3')

    Parameters
    ----------
    index : str
        Excel index

    Returns
    -------
    tuple coordinate index

    """
    [alpha, num, _] = re.split(r'(\d+)', index)
    return (int(num) - 1, alphaCol2Num(alpha))


def getAllSheetNames(filePath):
    """
    Retrive all sheetnames of the given excel file path.


    Parameters
    ----------
    filePath : str
        Path of excel file to get sheet names.

    Returns
    -------
    sheetNameList : list of str


    """
    xlFile = pd.ExcelFile(filePath)
    sheetNameList = xlFile.sheet_names
    xlFile.close()
    return sheetNameList


def getCountryExtract(countryList, sourceList='all'):
    """
    Extracts all available data in the database to an excel file
    """
    import datatoolbox as dt

    if sourceList == 'all':
        sourceList = list(dt.core.DB.sources.index)

    if not isinstance(countryList, list):
        countryList = [countryList]
    if not isinstance(sourceList, list):
        sourceList = [sourceList]

    #%%
    resFull = list()
    sourceList.sort()
    for source in tqdm.tqdm(sourceList):
        print(source)
        newList = list(dt.find(source=source).index)
        newList.sort()
        resFull = resFull + newList

    _res2Excel(resFull, countryList)


def pandasStr2floatPercent(X):
    return [_str2float(x) for x in X]


#%% Classes
def isnumber(x):
    try:
        float(x)
        return True
    except:
        return False
    
class Excel_Reader():
    
    def __init__(self, filename):
        self.filename = filename
        self.xls = pd.ExcelFile(self.filename, engine='calamine')

    
    def extract_data(self, 
                     xlsx_layout :dict,
                     data_layout: dict,
                     column_mapping: dict = None,
                     index_mapping: dict = None,
                     meta = None,
                     units= None,
                     ):
        
        #self = reader
        # xlsx_layout = test_config['xlsx_layout']
        # data_layout = test_config['data_layout']
        # column_mapping = test_config['column_mapping']
        # index_mapping = test_config['index_mapping']
        
        df = pd.read_excel(self.xls, **xlsx_layout)
        df.columns.names = data_layout['column_names']
        df.index.names = data_layout['index_names']
        
        df = df[df.map(isnumber)]
        
        missing_in_col =[x for x in column_mapping.keys() if not x in df.columns]
        if len(missing_in_col)>0:
            raise(Exception(
                f'{df.columns} does not contain: {missing_in_col}')
                )
        #apply mapping
        df.columns = df.columns.map(column_mapping, na_action='ignore')
       
        #% Process index names
        df.index = df.index.str.strip() # remove leading and tailing spaces
       
        missing_in_idx =[x for x in index_mapping.keys() if not x in df.index]
        if len(missing_in_idx)>0:
            raise(Exception(
                f'{df.index} does not contain: {missing_in_idx}')
                )
           
        #apply mapping
        df.index = df.index.map(index_mapping)
        df = df.loc[df.index[~df.index.isnull()], df.columns[~df.columns.isnull()]]

        
        ldf = pd.melt(df.reset_index(), id_vars=['category'], value_vars=df.columns)
        
        
        ldf['unit'] =  ldf['entity'].map(units)
        
        index_levels = data_layout["column_names"] + data_layout["index_names"] + ['unit']
        ldf = ldf.set_index(index_levels)
        if meta is not None :
            for meta_key, meta_value in meta.items():
                ldf = ldf.idx.assign(**{meta_key : meta_value})
        return ldf
    
    def openSourceFile(self):
        if OS == 'Linux':
            os.system('libreoffice ' + self.filename)
        elif OS == 'Darwin':
            os.system('open -a "Microsoft Excel" ' + self.filename)




class ExcelWriter:
    """
    More complex excel interface
    Author: Andreas Geiges
    """

    def __init__(
        self,
        setup=None,
        fileName=None,
        overwrite=False,
        setupSheet='INPUT',
        interpolate=False,
    ):
        self.overwrite = overwrite  # overwriting values
        self.setupSheet = setupSheet
        if setup:
            self.setup = setup
            fileSetup = True
        else:
            self.setup = dict()
            self.setup['fileName'] = fileName
            fileSetup = False

        self.interpolate = interpolate

    def close(self):
        self.wb.close()

    def getSetups(self):
        self.wb_read = load_workbook(self.setup['fileName'], data_only=True)
        self.wb = load_workbook(self.setup['fileName'])
        setup = dict()
        setup['fileName'] = self.setup['fileName']

        # new using pandas
        mapping = pd.read_excel(self.setup['fileName'], sheet_name=self.setupSheet)
        mapping.columns = [x.lower() for x in mapping.columns]

        setupDict = {
            'sheetName': 'sheetname',
            'timeIdxList': 'time',
            'spaceIdxList': 'region',
            'dataID': 'variable',
            'unit': 'unit',
            #                     'unitTo'        : 'unitto'
        }

        for i, setupMapp in mapping.iterrows():
            print(setupMapp)
            for key in setupDict.keys():
                setup[key] = str(setupMapp[setupDict[key]])
                if setup[key] == 'nan':
                    setup[key] = None
            yield (dict(setup))
    

    def insert_data(self):
        replaceDict = {'EU': 'EU28'}


        # Copy old file with "_filled" extension and load it
        if '.xlsx' in self.setup['fileName']:
            saveFileName = self.setup['fileName'].replace('.xlsx', '_filled.xlsx')
        elif '.xls' in self.setup['fileName']:
            saveFileName = self.setup['fileName'].replace('.xls', '_filled.xlsx')
        else:
            print('file extention not recognized')
        copyfile(self.setup['fileName'], saveFileName)
        self.setup['fileName'] = saveFileName
        #        wb = load_workbook(self.setup['fileName'])

        # pre-load all valid spatial IDS
        validSpatialIDs = mapp.getValidSpatialIDs()

        # create list of setups
        self.setupList = list()

        # loop over all defined inputs in the INPUT sheet mapping (rows)
        for setup in self.getSetups():

            # ensure input is at all times a string
            if isinstance(setup['timeIdxList'], int):
                setup['timeIdxList'] = str(setup['timeIdxList'])
            if isinstance(setup['spaceIdxList'], int):
                setup['spaceIdxList'] = str(setup['spaceIdxList'])

            if config.DEBUG:
                print(setup)

            # add a copy of setup to list
            self.setupList.append(copy(setup))

            args = None, None, None, None, None
            wksSheet_read = self.wb_read[setup['sheetName']]
            wksSheet_write = self.wb[setup['sheetName']]
            # loop overall setups and collect dataIDs for pre-loading
            tableIDs = list()
            for args in _iterData(setup['dataID'], wksSheet_read, *args):

                if config.DEBUG:
                    print(args)
                if (~pd.isna(args[4])) and (args[4] is not None):
                    tableIDs.append(args[4])

            if config.DEBUG:
                print(tableIDs)

            # load all tables from database
            tables = core.DB.getTables(tableIDs)

            if setup['unit'] is not None:
                for tableKey in tables.keys():
                    table = tables[tableKey]
                    tables[tableKey] = table.convert(setup['unit'])

            if self.interpolate:
                for tableKey in tables.keys():
                    for col in list(
                        range(
                            tables[tableKey].columns.min(),
                            tables[tableKey].columns.max() + 1,
                        )
                    ):
                        if col not in tables[tableKey].columns:
                            tables[tableKey].loc[:, col] = np.nan
                    tables[tableKey] = tables[tableKey].interpolate()

            #            wksSheet = load_workbook(setup['fileName'], data_only=True)[setup['sheetName']]

            args = [None, None, None, None, None]
            iCount = 0

            # iterate over time index list
            for argsTime in _iterTime(setup['timeIdxList'], wksSheet_read, *args):
                if argsTime[TI_ARG] is None:
                    print('No time defintion found')
                    continue

                # iterate over space index list
                for argsSpace in _iterSpace(
                    setup['spaceIdxList'], wksSheet_read, *argsTime
                ):
                    if argsSpace[SP_ARG] is None:
                        print('not spacial defintion found')
                        continue

                    if argsSpace[SP_ARG] not in validSpatialIDs:
                        #                        print('not time defintion found')
                        #                        print(argsSpace[SP_ARG])
                        if argsSpace[SP_ARG] in replaceDict.keys():
                            newID = replaceDict[argsSpace[SP_ARG]]
                        else:
                            newID = util.getCountryISO(argsSpace[SP_ARG])

                        if newID in validSpatialIDs:
                            argsSpace[SP_ARG] = newID
                            print(argsSpace[SP_ARG])
                        else:
                            print(argsSpace[SP_ARG] + ' not found')
                            continue

                    # print(argsSpace[SP_ARG])
                    # iterate over all data indices
                    for argData in _iterData(
                        setup['dataID'], wksSheet_read, *argsSpace
                    ):
                        if argData[DT_ARG] is None:
                            print('not data defintion found')
                            continue

                        try:
                            value = tables[argData[DT_ARG]].loc[
                                argData[SP_ARG], int(argData[TI_ARG])
                            ]
                            self._writeValueNew(
                                wksSheet_write,
                                xlsRow=argData[0],
                                xlsCol=argData[1],
                                value=value,
                            )

                            #                            print('success')
                            iCount += 1

                        except Exception as e:
                            if config.DEBUG:
                                print('failed with agruments:' + str(argData))
                                print(e)
                                # import pdb
                                # pdb.set_trace()
                            pass

            self.wb.save(self.setup['fileName'])
            print(setup)
            print('{} items inserted'.format(iCount))

            self.wb.close()

    def _writeValueNew(self, wrkSheet, xlsRow, xlsCol, value):
        if self.overwrite or pd.isna(wrkSheet.cell(row=xlsRow, column=xlsCol).value):
            print(value)
            wrkSheet.cell(row=xlsRow, column=xlsCol, value=value)


#    def _writeMultipleIndicators(self, setup):
#        ws_readValues = load_workbook(self.setup['fileName'], data_only=True)[setup['sheetName']]
#        try:
#            tableIDs = [cell.value for cell in ws_readValues[setup['dataID']]]
#        except:
#            tableIDs = [cell[0].value for cell in ws_readValues[setup['dataID']]]
#
#            tables = dt.getTables(tableIDs)
#
#            REG_FIND_ROWS.search(setup['timeIdxList'])


def createExcelInventory():
    """
    Create a excel inventory of the current data base. The excel files will be
    placed in the directory of the database (see config.PATH_TO_DATASHELF)

    Returns
    -------
    None.

    """
    #%%
    import datatoolbox as dt
    import xlwings as xl
    import pandas as pd
    import time
    import os
    import tqdm
    import math

    xl.books.open(dt.config.PATH_TO_DATASHELF + '/excel_inventory/main.xlsx')
    try:
        wb = xl.books.open(dt.config.PATH_TO_DATASHELF + '/excel_inventory/main.xlsx')
    except:
        wb = xl.books.add()
    try:
        sht = wb.sheets.add('Sources')
    except:
        sht = wb.sheets['Sources']
    row = 6
    sheets = dict()
    sht.range('B2').value = 'Datatoolbox Inventory - ' + time.strftime("%Y/%m/%d")

    sht.range('B5').value = ['Source ID', 'Licence', 'Collected by', 'Url']
    sht.range('B5:E5').color = (0, 104, 204)
    #    sht.range('B2').value
    for sourceID, source_data in tqdm.tqdm(
        dt.sourceInfo().loc['SDG_DB_2019':, :].iterrows()
    ):

        sht.range('B' + str(row)).add_hyperlink('#{}!A1'.format(sourceID), sourceID)
        sht.range('C' + str(row)).value = source_data.loc[
            ['licence', 'collected_by', 'source_url']
        ].values
        try:
            sheets[sourceID] = wb.sheets.add(sourceID, after=1)
        except:
            sheets[sourceID] = wb.sheets[sourceID]
        sheets[sourceID].range('A1').add_hyperlink('#Sources!A1', 'back to Index')
        sheets[sourceID].range('B2').value = sourceID + ' - Inventory'

        path = dt.config.PATH_TO_DATASHELF + 'excel_inventory/sources/'
        file_name = sourceID + '.xlsx'
        link = 'file://' + os.path.join(path + file_name) + '#Inventory!A1'

        sheets[sourceID].range('D2').add_hyperlink(link, 'detailed inventory...')

        data_source = dt.find(source=sourceID)
        #        pathways = list(data_source.pathway.unique())
        entities = sorted(list(data_source.variable.unique()))

        sheets[sourceID].range((5, 2)).value = [
            'Variable',
            'Entity',
            'Category',
            'Historic',
            'Modeled',
        ]
        sheets[sourceID].range((5, 2), (5, 6)).color = (0, 102, 204)

        data_to_excel = pd.DataFrame(
            columns=['Variable', 'Entity', 'Category', 'Historic', 'Modeled']
        )
        for index, entity in enumerate(entities):
            #
            ##            categories = sorted(list(dt.findp(source=sourceID).category.unique()))
            mask = data_source.variable == entity
            data = data_source.loc[data_source.index[mask], :]
            data_to_excel.loc[index, ['Variable', 'Entity', 'Category']] = data.loc[
                data.index[0], ['variable', 'entity', 'category']
            ].values

            hist_scen = [
                x for x in data.pathway.unique() if ('Hist' in x) or ('hist' in x)
            ]
            data_to_excel.loc[index, ['Historic']] = '{} hist pathways'.format(
                len(hist_scen)
            )

            other_scen = [
                x for x in data.pathway.unique() if ('Hist' in x) or ('hist' in x)
            ]
            data_to_excel.loc[index, ['Modeled']] = '{} other pathways'.format(
                len(data.pathway.unique()) - len(hist_scen)
            )

        sheets[sourceID].range((6, 2)).value = data_to_excel.values
        sheets[sourceID].range((6, 2), (30, 6)).autofit()

        if row % 2 == 0:
            sht.range('B{0}:E{0}'.format(row)).color = (153, 204, 255)
        else:
            sht.range('B{0}:E{0}'.format(row)).color = (204, 229, 255)

        row += 1
        #        dsf
        _create_source_inventory(sourceID)

    sht.autofit()



def _create_source_inventory(sourceID):
    #%
    import os
    import datatoolbox as dt
    import xlwings as xl
    import pandas as pd
    import time

    path = dt.config.PATH_TO_DATASHELF + '/excel_inventory/sources/'
    file_name = sourceID + '.xlsx'
    try:
        wb = xl.books.open(os.path.join(path, file_name))
    except:
        wb = xl.books.add()
        wb.save(os.path.join(path, file_name))
    try:
        sht = wb.sheets.add('Inventory')
    except:
        sht = wb.sheets['Inventory']
    sheets = dict()
    path_link = dt.config.PATH_TO_DATASHELF + 'excel_inventory/'
    file_name_link = 'main.xlsx'
    link = (
        'file://' + os.path.join(path_link + file_name_link) + '#{}!A1'.format(sourceID)
    )

    sht.range('A1').add_hyperlink(link, 'back to Main inventory')
    sht.range('B2').value = sourceID
    sht.range('D2').value = ['Created:', time.strftime("%Y/%m/%d")]
    sht.range('D3').value = ['Licence:', dt.sourceInfo().loc[sourceID, 'licence']]

    sht.range('B5').value = 'Pathways:'
    sht.range('B5:G5').color = (0, 104, 208)
    data_source = dt.find(source=sourceID)
    row = 6
    for pathway in sorted(list(data_source.pathway.unique())):
        excel_pathway_string = (
            pathway.replace('|', '').replace('-', '_').replace('.', '')
        )
        sht.range('B' + str(row)).add_hyperlink(
            '#{}!A1'.format(excel_pathway_string[:30]), excel_pathway_string
        )
        if row % 2 == 0:
            sht.range('B{0}:G{0}'.format(row)).color = (153, 204, 255)
        else:
            sht.range('B{0}:G{0}'.format(row)).color = (204, 229, 255)

        try:
            sheets[excel_pathway_string] = wb.sheets.add(
                excel_pathway_string[:30], after=1
            )
        except:
            sheets[excel_pathway_string] = wb.sheets[excel_pathway_string[:30]]

        mask = data_source.pathway == pathway
        data = data_source.loc[data_source.index[mask], :]
        sheets[excel_pathway_string].range('B4').value = ['Entity', 'Category']
        sheets[excel_pathway_string].range('B5').value = (
            data.sort_values(by='variable').loc[:, ['entity', 'category']].values
        )
        sheets[excel_pathway_string].range('D4').value = data.sort_values(
            by='variable'
        ).loc[:, ['unit']]

        sheets[excel_pathway_string].autofit()
        sheets[excel_pathway_string].range('A1').add_hyperlink(
            '#Inventory!A1', 'back to Index'
        )
        sheets[excel_pathway_string].range('B2').value = [
            excel_pathway_string,
            ' - Inventory - ',
        ]
        sheets[excel_pathway_string].range('D4').value = 'Table ID'
        sheets[excel_pathway_string].range((4, 2), (4, 5)).color = (0, 102, 204)
        row += 1
    sht.autofit()
    wb.save(os.path.join(path, file_name))
    wb.close()
    # xl.Range('B1').add_hyperlink('file://Book2.xlsx#Sheet2!A1',"Sheet2")


# import xlwings as xw
# from xlwings import Book, Range
# from datetime import datetime
# import numpy as np
# wb = Book('Book2')
# Range('A2').value = 12
# Range('A1').expand('table').value
# Range('B1').value = [[1],[2],[3],[4],[5]]
# Range('B').value
#
# xw.Range('B2').value = 222
# xw.Range('A2').expand('table').value
# xw.Range('A2').expand('down').value
# xw.Range('A2').options(ndim=2).value
#
#
##%%
# app = xw.apps.active
# wb = app.books.active
# sht = wb.sheets.active
#
# used_range = sht.used_range.address
#
# Range(used_range).expand('table').value
#
# xw.Range('B1').add_hyperlink('[Book2]Sheet2!$A$1',"Sheet2")
# xw.Range('B1').add_hyperlink(':Sheet2.A1',"Sheet2")
# xw.Range('B1').add_hyperlink('file:///Users/andreasgeiges/Documents/Book2.xlsx#Sheet2!A1',"Sheet2")
#


def read_active_excel_sheet():
    """
    Read the entire active excel sheet to a pandas dataframe.

    Return pandas.Dataframe
    """
    import xlwings as xw

    app = xw.apps.active
    book = app.books.active
    sheet = book.sheets[0]
    df = pd.DataFrame(sheet.used_range.value)
    return df
